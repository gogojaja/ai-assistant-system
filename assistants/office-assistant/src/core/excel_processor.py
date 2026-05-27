#!/usr/bin/env python3

"""
模块名称：excel_processor
功能描述：Excel 文件处理模块（结构分析、数据提取）
对外接口：
    - ExcelProcessor(file_path): 初始化并分析 Excel
        - analyze(): 返回结构摘要
        - get_data_text(): 提取表格文本供模型使用
依赖：
    - 标准库：logging
    - 第三方：openpyxl
    - 项目内：无
版本：v1.0
更新记录：
    - 2026-05-23: 添加统一注释头，支持智能摘要数据提取
"""
import logging
import openpyxl

logger = logging.getLogger(__name__)


class ExcelProcessor:
    MAX_ROWS_FOR_TEXT = 20
    MAX_COLS_FOR_TEXT = 15

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self._load()

    def _load(self):
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        except Exception as e:
            logger.error(f"无法打开 Excel 文件: {e}")
            raise

    def analyze(self) -> str:
        """简洁结构摘要（不含具体首行文字，节省空间）"""
        if not self.workbook:
            return "Excel 文件未能正确加载。"
        try:
            summary_parts = [
                f"📊 工作簿包含 {len(self.workbook.sheetnames)} 个工作表："
                f"{', '.join(self.workbook.sheetnames)}"
            ]
            total_rows = 0
            for name in self.workbook.sheetnames:
                ws = self.workbook[name]
                rows = ws.max_row
                cols = ws.max_column
                total_rows += rows
                summary_parts.append(f"· {name}: {rows}行×{cols}列")
            summary_parts.append(f"📈 总计约 {total_rows} 行数据")
            return '\n'.join(summary_parts)
        except Exception as e:
            logger.error(f"Excel分析失败: {e}")
            return "Excel 文件解析失败，请确认文件格式正确。"

    def get_data_text(self) -> str:
        """
        提取每个工作表前 N 行和列，格式化为可供大模型阅读的文本。
        增加数据量，帮助模型生成有意义摘要。
        """
        if not self.workbook:
            return ""
        text_parts = []
        for name in self.workbook.sheetnames:
            ws = self.workbook[name]
            max_row = min(ws.max_row, self.MAX_ROWS_FOR_TEXT)
            max_col = min(ws.max_column, self.MAX_COLS_FOR_TEXT)
            text_parts.append(f"【工作表：{name}】")
            for row_idx in range(1, max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell_val = ws.cell(row_idx, col_idx).value
                    if cell_val is None:
                        cell_val = ""
                    cell_str = str(cell_val).replace('\n', ' ').replace('\r', ' ').strip()
                    row_data.append(cell_str)
                text_parts.append('\t'.join(row_data))
            text_parts.append("")
        return '\n'.join(text_parts)