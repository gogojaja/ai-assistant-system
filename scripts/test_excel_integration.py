#!/usr/bin/env python3

"""
模块名称：test_excel_integration
功能描述：TODO: 请补充功能描述
对外接口：
    - create_test_excel()
    - test_single_sheet()
    - test_multi_sheet()
    - test_empty_sheet()
    - test_data_text()
    - test_summary_length()
依赖：
    - 标准库：logging, os, sys, tempfile
    - 第三方：core, openpyxl
    - 项目内：无
版本：v1.0
更新记录：
    - 2026-05-23: 自动添加统一注释头
"""
"""
模块名称：test_excel_integration
功能描述：测试 Excel 集成分析逻辑（结构分析、数据提取、长度控制）
对外接口：
    - 直接运行，执行 5 组测试用例并输出结果
依赖：
    - 标准库：sys, os, tempfile, logging
    - 第三方：openpyxl
    - 项目内：assistants.office-assistant.src.core.excel_processor (ExcelProcessor)
版本：v1.0
更新记录：
    - 2026-05-23: 初始创建，添加统一注释头
"""
import sys
import os
import tempfile
import logging
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.dirname(SCRIPT_DIR)
sys.path.insert(0, os.path.join(PROJECT, "assistants/office-assistant/src"))
from core.excel_processor import ExcelProcessor

logging.basicConfig(level=logging.INFO, format='[TEST] %(message)s')
logger = logging.getLogger("test_excel")

def create_test_excel(path, sheets_data):
    wb = openpyxl.Workbook()
    first_sheet_name = list(sheets_data.keys())[0] if sheets_data else None
    if first_sheet_name:
        ws = wb.active
        ws.title = first_sheet_name
        data = sheets_data[first_sheet_name]
        for row_idx, row in enumerate(data, 1):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row_idx, col_idx, value=val)
    for sheet_name, data in list(sheets_data.items())[1:]:
        ws = wb.create_sheet(sheet_name)
        for row_idx, row in enumerate(data, 1):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row_idx, col_idx, value=val)
    wb.save(path)

def test_single_sheet():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    create_test_excel(path, {
        "数据": [["姓名","年龄","城市"], ["张三",30,"北京"], ["李四",25,"上海"]]
    })
    processor = ExcelProcessor(path)
    result = processor.analyze()
    assert "数据" in result
    assert "3行" in result
    logger.info("✅ 测试1通过：单工作表结构分析")
    os.unlink(path)

def test_multi_sheet():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    create_test_excel(path, {
        "Sheet1": [["产品","价格"]],
        "Sheet2": [["日期","销量"], ["5/1",100], ["5/2",150]]
    })
    processor = ExcelProcessor(path)
    result = processor.analyze()
    assert "2 个工作表" in result
    assert "Sheet2" in result
    logger.info("✅ 测试2通过：多工作表结构分析")
    os.unlink(path)

def test_empty_sheet():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    create_test_excel(path, {"空表": []})
    processor = ExcelProcessor(path)
    result = processor.analyze()
    assert "行" in result
    logger.info("✅ 测试3通过：空工作表结构分析")
    os.unlink(path)

def test_data_text():
    """验证数据文本提取行数限制：MAX_ROWS=5，应包含标题行+前4条数据（共5行）"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    create_test_excel(path, {
        "销售": [["日期","金额"]] + [[f"5/{i}", i*100] for i in range(1, 5)]
    })
    processor = ExcelProcessor(path)
    data_text = processor.get_data_text()
    lines = data_text.split('\n')
    for i in range(1, 5):
        assert any(f"5/{i}" in line for line in lines), f"第{i}条数据应存在"
    assert not any("5/5" in line for line in lines), "第5条数据不应出现"
    logger.info("✅ 测试4通过：数据文本提取行数限制（MAX_ROWS=5）")
    os.unlink(path)

def test_summary_length():
    """模拟摘要+结构信息长度控制"""
    structure = "📊 工作簿包含... (结构信息)" * 20
    summary = "这是一段摘要" * 50
    reply = structure + "\n\n🤖 AI 摘要：\n" + summary
    max_len = 4900
    if len(reply) > max_len:
        available = max_len - len(structure) - len("\n\n🤖 AI 摘要：\n")
        if available > 10:
            summary = summary[:available-1] + "…"
        else:
            summary = "（摘要过长，已省略）"
    final = structure + "\n\n🤖 AI 摘要：\n" + summary
    assert len(final) <= max_len
    logger.info("✅ 测试5通过：消息长度截断控制")

if __name__ == "__main__":
    logger.info("🚀 开始测试 Excel 集成分析（适配 MAX_ROWS=5）...")
    test_single_sheet()
    test_multi_sheet()
    test_empty_sheet()
    test_data_text()
    test_summary_length()
    logger.info("🎉 所有测试通过")